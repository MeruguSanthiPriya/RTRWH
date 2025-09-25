import pandas as pd
from math import radians, sin, cos, sqrt, asin
from flask import Flask, request, render_template, redirect, url_for, jsonify, send_from_directory, make_response, session, flash
from flask_cors import CORS
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import os
import openpyxl
from fpdf import FPDF, XPos, YPos
from datetime import datetime
import bcrypt
from functools import wraps
from recommendations import determine_category, calculate_structure_dimensions, estimate_costs_and_payback, get_purification_recommendations, calculate_harvesting_potential
import requests
from database import db
from models import AquiferMaterial # Add other models as you create them
from geoalchemy2 import WKTElement
from sqlalchemy import func
import json

# --- Validation Functions ---
def validate_name(name):
    """Validate name is not empty and has reasonable length"""
    if not name or not name.strip():
        return False, "Name cannot be empty"
    if len(name.strip()) < 2:
        return False, "Name must be at least 2 characters long"
    if len(name.strip()) > 100:
        return False, "Name must be less than 100 characters long"
    return True, name.strip()

def validate_location_name(location_name):
    """Validate location name is not empty and has reasonable length"""
    if not location_name or not location_name.strip():
        return False, "Location name cannot be empty"
    if len(location_name.strip()) < 2:
        return False, "Location name must be at least 2 characters long"
    if len(location_name.strip()) > 200:
        return False, "Location name must be less than 200 characters long"
    return True, location_name.strip()

def validate_latitude(lat):
    """Validate latitude is within valid range (-90 to 90)"""
    try:
        lat_float = float(lat)
        if not (-90 <= lat_float <= 90):
            return False, "Latitude must be between -90 and 90 degrees"
        return True, lat_float
    except (ValueError, TypeError):
        return False, "Invalid latitude format"

def validate_longitude(lon):
    """Validate longitude is within valid range (-180 to 180)"""
    try:
        lon_float = float(lon)
        if not (-180 <= lon_float <= 180):
            return False, "Longitude must be between -180 and 180 degrees"
        return True, lon_float
    except (ValueError, TypeError):
        return False, "Invalid longitude format"

def validate_rooftop_area(area):
    """Validate rooftop area is within reasonable bounds (1-5000 sq.m)"""
    try:
        area_float = float(area)
        if not (1 <= area_float <= 5000):
            return False, "Rooftop area must be between 1 and 5,000 square meters"
        return True, area_float
    except (ValueError, TypeError):
        return False, "Invalid rooftop area format"

def validate_open_space_area(area):
    """Validate open space area is within reasonable bounds (0-10000 sq.m)"""
    try:
        area_float = float(area)
        if not (0 <= area_float <= 10000):
            return False, "Open space area must be between 0 and 10,000 square meters"
        return True, area_float
    except (ValueError, TypeError):
        return False, "Invalid open space area format"

def validate_household_size(size):
    """Validate household size is within reasonable bounds (1-50 people)"""
    try:
        size_int = int(size)
        if not (1 <= size_int <= 50):
            return False, "Household size must be between 1 and 50 people"
        return True, size_int
    except (ValueError, TypeError):
        return False, "Invalid household size format"

# Initialize the Flask app
app = Flask(__name__)
CORS(app)  # Allow cross-origin requests from frontend

# Configure the secret key for sessions
app.config['SECRET_KEY'] = 'your-secret-key-change-in-production'  # Change this in production

# Configure the database file
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://app_user:password@localhost:5432/rtrwh_gis'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize the database with the app
db.init_app(app)

# --- One-time lightweight schema guard for newly added fields ---
def _ensure_user_input_new_columns():
    """Safely add newly introduced columns to user_input table if they don't exist.

    This avoids immediate migration tooling overhead. For production
    replace with proper Alembic migration.
    """
    from sqlalchemy import text
    ddl_statements = [
        "ALTER TABLE user_input ADD COLUMN IF NOT EXISTS building_age VARCHAR(30)",
        "ALTER TABLE user_input ADD COLUMN IF NOT EXISTS occupancy INTEGER",
        # Safe attempt to drop legacy column if present
        "DO $$ BEGIN IF EXISTS (SELECT 1 FROM information_schema.columns WHERE table_name='user_input' AND column_name='budget_preference') THEN ALTER TABLE user_input DROP COLUMN budget_preference; END IF; END $$;"
    ]
    for ddl in ddl_statements:
        try:
            db.session.execute(text(ddl))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"[Schema Guard] Skipped/failed: {ddl} -> {e}")

with app.app_context():
    _ensure_user_input_new_columns()

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'admin_login'
login_manager.login_message = 'Please log in to access the admin panel.'

# --- Path Configuration ---
BASE_DIR = os.path.abspath(os.path.dirname(__file__))

# --- Pre-load Data ---
# Location data is not available - using database queries instead
location_df = None

# --- Derived Categories from CSV ---
def derive_region_categories(df):
    """Analyze the CSV and assign a recommended RWH category per region.

    Returns a DataFrame with Region_Name, State and assigned category info.
    """
    if df is None or df.empty:
        return None

    def assign_category_from_row(row):    
        # Use the centralized `determine_category` function for consistent logic
        category_info = determine_category(
            roof_area=100,  # Assume a standard medium-sized roof for regional summary
            open_space=50, # Assume a standard medium-sized open space
            rainfall=float(row.get('Rainfall_mm', 0)),
            soil_type=str(row.get('Soil_Type', 'Loamy')),
            gw_depth=float(row.get('Groundwater_Depth_m', 10)),
            infiltration_rate=float(row.get('Infiltration_Rate_mm_per_hr', 15))
        )
        return category_info['category']
        
    df_copy = df.copy()
    df_copy['RWH_Category'] = df_copy.apply(assign_category_from_row, axis=1)
    # Map category to human-friendly name
    cat_map = {
        1: 'Storage Tank Only',
        2: 'Storage + Small Recharge Pit',
        3: 'Recharge Pit/Trench + Storage Tank',
        4: 'Recharge Shaft / Borewell Recharge',
        5: 'Recharge Pond / Community Structures',
        6: 'Supplementary Only'
    }
    df_copy['RWH_Category_Name'] = df_copy['RWH_Category'].map(cat_map)
    return df_copy[['Region_Name', 'State', 'Latitude', 'Longitude', 'Rainfall_mm', 'Soil_Type', 'Aquifer_Type', 'Infiltration_Rate_mm_per_hr', 'Groundwater_Depth_m', 'RWH_Category', 'RWH_Category_Name']]


@app.route('/api/regions_categories')
def api_regions_categories():
    """Return region-level RWH category assignment derived from the CSV."""
    if location_df is None:
        return jsonify({'error': 'Location CSV not available'}), 500

    df_categories = derive_region_categories(location_df)
    if df_categories is None:
        return jsonify({'error': 'No data'}), 404

    records = df_categories.to_dict(orient='records')
    return jsonify({'regions': records, 'count': len(records)})


@app.route('/api/categories')
def api_categories():
    """Return summary counts per RWH category across the CSV."""
    if location_df is None:
        return jsonify({'error': 'Location CSV not available'}), 500
    df_categories = derive_region_categories(location_df)
    if df_categories is None:
        return jsonify({'error': 'No data'}), 404

    counts = df_categories['RWH_Category_Name'].value_counts().to_dict()
    # Provide a stable ordering
    ordered = {k: counts.get(k, 0) for k in ['Storage Tank Only', 'Storage + Small Recharge Pit', 'Recharge Pit/Trench + Storage Tank', 'Recharge Shaft / Borewell Recharge', 'Recharge Pond / Community Structures', 'Supplementary Only']}
    return jsonify({'category_counts': ordered})

# --- Database Model for User Data ---
class UserInput(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), nullable=False)
    location_name = db.Column(db.String(120), nullable=False)
    user_lat = db.Column(db.Float)
    user_lon = db.Column(db.Float)
    household_size = db.Column(db.Integer)
    rooftop_area = db.Column(db.Float)
    open_space_area = db.Column(db.Float)  # NEW FIELD
    roof_type = db.Column(db.String(50))
    property_type = db.Column(db.String(50))  # NEW FIELD
    existing_water_sources = db.Column(db.String(200))  # NEW FIELD
    # budget_preference = db.Column(db.String(50))  # REMOVED - budget option no longer used
    intended_use = db.Column(db.String(100))  # NEW FIELD
    building_age = db.Column(db.String(30))  # NEW FIELD: 'new' or 'existing'
    occupancy = db.Column(db.Integer)  # NEW FIELD: commercial/institutional occupancy
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

# --- Admin User Model ---
class AdminUser(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    role = db.Column(db.String(50), default='admin')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)
    
    def set_password(self, password):
        """Hash and set the password"""
        password_bytes = password.encode('utf-8')
        self.password_hash = bcrypt.hashpw(password_bytes, bcrypt.gensalt()).decode('utf-8')
    
    def check_password(self, password):
        """Check if the provided password matches the hash"""
        password_bytes = password.encode('utf-8')
        return bcrypt.checkpw(password_bytes, self.password_hash.encode('utf-8'))

# --- Geological Data Model ---
class GeoData(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    region_name = db.Column(db.String(120), nullable=False)
    state = db.Column(db.String(120), nullable=False)
    latitude = db.Column(db.Float, nullable=False)
    longitude = db.Column(db.Float, nullable=False)
    rainfall_mm = db.Column(db.Float)
    groundwater_depth_m = db.Column(db.Float)
    aquifer_type = db.Column(db.String(80))
    aquifer_depth_min_m = db.Column(db.Float)
    aquifer_depth_max_m = db.Column(db.Float)
    aquifer_thickness_m = db.Column(db.Float)
    remarks = db.Column(db.Text)
    soil_type = db.Column(db.String(80))
    infiltration_rate_mm_per_hr = db.Column(db.Float)
    soil_permability_class = db.Column(db.String(50))
    water_quality = db.Column(db.String(80))
    water_cost_per_liter = db.Column(db.Float, default=0.16)

    def to_dict(self):
        """Converts the object to a dictionary."""
        return {c.name: getattr(self, c.name) for c in self.__table__.columns}

# Flask-Login user loader
@login_manager.user_loader
def load_user(user_id):
    return AdminUser.query.get(int(user_id))

# Admin required decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

# --- Core Calculation Functions ---

def haversine(lat1, lon1, lat2, lon2):
    """Calculate the distance between two points on Earth using the Haversine formula."""
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a))
    r = 6371  # Radius of Earth in kilometers
    return c * r

def get_nearest_location(user_lat, user_lon):
    """Find the nearest location from the database based on user's GPS coordinates."""
    try:
        # Query ground water level stations for nearest location
        from models import GroundWaterLevelStation
        stations = GroundWaterLevelStation.query.all()

        if not stations:
            return None

        min_distance = float('inf')
        nearest_station = None

        for station in stations:
            if station.lat and station.long:
                distance = haversine(user_lat, user_lon, station.lat, station.long)
                if distance < min_distance:
                    min_distance = distance
                    nearest_station = station

        if nearest_station:
            return {
                'Region_Name': nearest_station.station_na or 'Unknown',
                'State': nearest_station.state_name or 'Unknown',
                'Latitude': nearest_station.lat,
                'Longitude': nearest_station.long,
                'distance': min_distance,
                'station_code': nearest_station.station_co,
                'agency': nearest_station.agency_nam,
                'district': nearest_station.district_n,
                'basin': nearest_station.basin_name
            }
    except Exception as e:
        print(f"Error finding nearest location: {e}")

    return None

def get_mock_location_data(location_name, user_lat=None, user_lon=None):
    """Get location data by name from the database."""
    try:
        from models import GroundWaterLevelStation
        # Try to find a station with a matching name
        stations = GroundWaterLevelStation.query.filter(
            GroundWaterLevelStation.station_na.ilike(f'%{location_name}%')
        ).limit(5).all()

        if stations:
            station = stations[0]  # Take the first match
            result = {
                'Region_Name': station.station_na or location_name,
                'State': station.state_name or 'Unknown',
                'Latitude': station.lat,
                'Longitude': station.long,
                'station_code': station.station_co,
                'agency': station.agency_nam,
                'district': station.district_n,
                'basin': station.basin_name
            }
            if user_lat and user_lon:
                result['distance'] = haversine(user_lat, user_lon, station.lat, station.long)
            return result

        # If no match found, return a default location or None
        return None
    except Exception as e:
        print(f"Error getting mock location data: {e}")
        return None

def calculate_runoff_potential(roof_area_m2, rainfall_mm, runoff_coefficient):
    """Calculate annual runoff generation capacity."""
    annual_runoff_liters = roof_area_m2 * rainfall_mm * runoff_coefficient
    peak_monthly = annual_runoff_liters * 0.4  # Assuming 40% in peak monsoon month
    return {
        'annual_liters': annual_runoff_liters,
        'peak_monthly': peak_monthly,
        'daily_average': annual_runoff_liters / 365
    }

def validate_artificial_recharge_safety(location_data):
    """Check if artificial recharge is safe based on multiple factors."""
    safety_issues = []
    is_safe = True
    
    # Check groundwater depth
    gw_depth = location_data.get('Groundwater_Depth_m', 10)
    if gw_depth < 3:
        safety_issues.append("Shallow groundwater (<3m) - Risk of waterlogging and contamination")
        is_safe = False
    
    # Check water quality
    water_quality = location_data.get('Water_Quality', 'Good')
    if water_quality.lower() in ['poor', 'contaminated']:
        safety_issues.append("Poor groundwater quality - Recharge may worsen contamination")
        is_safe = False
    
    # Check soil infiltration rate
    infiltration_rate = location_data.get('Infiltration_Rate_mm_per_hr', 15)
    if infiltration_rate < 5:
        safety_issues.append("Low soil infiltration (<5mm/hr) - Water will stagnate")
        is_safe = False
    
    # Check aquifer type and remarks for regulatory issues
    remarks = location_data.get('Remarks', '').lower()
    if 'overexploited' in remarks or 'prohibited' in remarks:
        safety_issues.append("Regulatory restrictions - Check CGWA guidelines")
        is_safe = False
    
    return {
        'is_safe': is_safe,
        'safety_issues': safety_issues,
        'alternatives': ['Storage tank only', 'Community structures', 'Water conservation'] if not is_safe else []
    }

def calculate_water_source_priority(existing_sources):
    """Calculate how much RWH system is needed based on existing sources reliability and cost"""
    if not existing_sources:
        return 0  # No existing sources = high priority for RWH

    priority_score = 0
    sources = [s.strip() for s in existing_sources.split(',')]

    # High priority sources (expensive/unreliable)
    if 'Water Tanker' in sources:
        priority_score += 40  # Very expensive, unreliable
    if 'Borewell' in sources:
        priority_score += 25  # Energy intensive, depleting groundwater
    if 'Open Well' in sources:
        priority_score += 20  # Contamination risk, seasonal availability

    # Medium priority sources
    if 'Municipal Supply' in sources:
        priority_score += 10  # Generally reliable but expensive

    # Low priority sources (good existing alternatives)
    if 'Private Borewell' in sources:
        priority_score += 5   # Better than public sources but still costly

    return min(priority_score, 50)  # Cap at 50 to avoid over-prioritization

def calculate_comprehensive_feasibility(location_data, user_input):
    """Enhanced feasibility calculation with safety checks and categorization."""

    # Extract parameters
    rainfall_mm = location_data['Rainfall_mm']
    roof_area = user_input.rooftop_area
    open_space = user_input.open_space_area or 0
    runoff_coeff = location_data.get('Runoff_Coefficient', 0.8)
    household_size = user_input.household_size
    soil_type = location_data.get('Soil_Type', 'Loamy')
    gw_depth = location_data.get('Groundwater_Depth_m', 10)
    infiltration_rate = location_data.get('Infiltration_Rate_mm_per_hr', 15)

    # Calculate runoff potential
    runoff_data = calculate_runoff_potential(roof_area, rainfall_mm, runoff_coeff)

    # Calculate harvesting potential with roof-type specific runoff coefficient
    harvesting_potential = calculate_harvesting_potential(roof_area, rainfall_mm, user_input.roof_type)

    # Check artificial recharge safety
    safety_check = validate_artificial_recharge_safety(location_data)

    # Calculate water source priority
    water_source_priority = calculate_water_source_priority(user_input.existing_water_sources)

    # Determine category with enhanced scoring
    user_preferences = {
        'complexity': 'balanced'  # Could be enhanced based on user input
    }

    category_result = determine_category(
        roof_area, open_space, rainfall_mm, soil_type, gw_depth, infiltration_rate,
        user_preferences,
        building_age=user_input.building_age,
        occupancy=user_input.occupancy,
        roof_type=user_input.roof_type,
        water_quality_required=(user_input.intended_use.lower() if user_input.intended_use else None),
        water_demand=(user_input.intended_use.lower() if user_input.intended_use else None),
        modification_type='retrofit' if (user_input.building_age in ['old', 'heritage']) else None,
        space_constraints='limited' if open_space < 20 else None,
        building_type=user_input.property_type
    )

    # Extract primary category for backward compatibility
    category_info = {
        'category': category_result['primary']['category'].category_id,
        'name': category_result['primary']['category'].name,
        'description': category_result['primary']['category'].description,
        'recommended_structures': category_result['primary']['category'].recommended_structures,
        'recharge_feasible': category_result['primary']['category'].recharge_feasible,
        'confidence_score': category_result['primary']['confidence'],
        'recommendation_reason': category_result['primary']['recommendation_reason'],
        'alternative_categories': category_result['alternatives']
    }

    # Calculate structure dimensions
    structure_dims = calculate_structure_dimensions(
        harvesting_potential['annual_liters'],
        infiltration_rate,
        open_space,
        category_info['recharge_feasible']
    )

    # Estimate costs and payback
    local_water_cost = location_data.get('Water_Cost_per_Liter', 0.16) # Fallback to 0.16 if not in CSV
    location_type = location_data.get('Location_Type', 'urban')  # Default to urban
    soil_type = location_data.get('Soil_Type', 'loamy')  # Default to loamy

    cost_analysis = estimate_costs_and_payback(
        category_id=category_info['category'],
        location_type=location_type,
        soil_type=soil_type,
        system_size=harvesting_potential['annual_liters'],
        intended_use=user_input.intended_use
    )

    # Get purification recommendations
    purification = get_purification_recommendations(
        user_input.intended_use or 'general',
        user_input.roof_type,
        location_data
    )

    # Calculate household demand
    daily_demand = household_size * 135  # liters per day
    annual_demand = daily_demand * 365

    # Overall feasibility score (adjusted by water source priority)
    if annual_demand > 0:
        base_feasibility = min((harvesting_potential['annual_liters'] / annual_demand) * 100, 100)
        # Apply water source priority boost (up to 20% increase for high-priority cases)
        priority_boost = min(water_source_priority * 0.4, 20)  # Max 20% boost
        feasibility_percentage = min(base_feasibility + priority_boost, 100)
    else:
        # If there is no demand (e.g., household size is 0), feasibility is not applicable.
        feasibility_percentage = 0.0

    if feasibility_percentage >= 80:
        feasibility_status = "Fully Feasible"
    elif feasibility_percentage >= 50:
        feasibility_status = "Partially Feasible"
    elif feasibility_percentage >= 20:
        feasibility_status = "Limited Feasible"
    else:
        feasibility_status = "Not Feasible"

    return {
        'runoff_data': runoff_data,
        'harvesting_potential': harvesting_potential,
        'safety_check': safety_check,
        'water_source_priority': water_source_priority,
        'category': category_info,
        'structure_dimensions': structure_dims,
        'cost_analysis': cost_analysis,
        'purification': purification,
        'annual_demand': annual_demand,
        'feasibility_percentage': round(feasibility_percentage, 1),
        'feasibility_status': feasibility_status
    }

# --- Flask Routes ---

@app.route('/')
def index_page():
    """Serves the main index.html page."""
    return render_template('index.html')

@app.route('/location-input')
def location_input_page():
    """Serves the location input page."""
    return render_template('location-input.html')

@app.route('/subsidy-checker.html')
def subsidy_checker_page():
    """Serves the static subsidy checker HTML file."""
    return render_template('subsidy-checker.html')

@app.route('/resources')
def resources_page():
    """Serves the resources page with learning materials."""
    return render_template('resources.html')

@app.route('/submit_location', methods=['POST'])
def submit_location():
    """
    Handles the initial location submission from the map.
    Stores location data in the session and shows the assessment type choice page.
    """
    data = request.get_json()
    lat = data.get('lat')
    lon = data.get('lon')
    address = data.get('address')

    # Basic validation
    if not lat or not lon or not address:
        return jsonify({'error': 'Missing location data'}), 400

    # Store location data in session
    session['latitude'] = lat
    session['longitude'] = lon
    session['address'] = address
    
    # Instead of redirecting, we'll confirm success and let the frontend handle the redirect
    return jsonify({'message': 'Location received, proceed to assessment type selection.', 'redirect_url': url_for('select_assessment_page')})

@app.route('/assessment-type')
def select_assessment_page():
    """Renders the page for choosing between Individual and Community assessment."""
    # Ensure location data is in session before showing this page
    if 'latitude' not in session:
        flash('Please select a location on the map first.', 'warning')
        return redirect(url_for('location_input_page'))
    return render_template('assessment-type.html')

@app.route('/select_assessment', methods=['POST'])
def select_assessment():
    """
    Handles the assessment type choice and redirects to the
    appropriate detailed input form.
    """
    assessment_type = request.form.get('assessment_type')

    if not assessment_type:
        flash('Please select an assessment type.', 'error')
        return redirect(url_for('select_assessment_page'))

    # Store the choice in the session
    session['assessment_type'] = assessment_type

    if assessment_type == 'Individual':
        # Redirect to a new page for individual inputs
        return redirect(url_for('individual_input_page'))
    elif assessment_type == 'Community':
        # Redirect to the existing community input page
        return redirect(url_for('community_input_page'))
    else:
        flash('Invalid assessment type selected.', 'error')
        return redirect(url_for('select_assessment_page'))

@app.route('/individual-input')
def individual_input_page():
    """Serves the page for individual user inputs, after location is set."""
    if 'latitude' not in session:
        flash('Please select a location first.', 'warning')
        return redirect(url_for('location_input_page'))
    return render_template('individual-input.html')


@app.route('/submit_form', methods=['POST'])
def submit_form():
    # --- Retrieve location data from session ---
    user_lat = session.get('latitude')
    user_lon = session.get('longitude')
    location_name = session.get('address')

    if not all([user_lat, user_lon, location_name]):
        flash("Your session has expired. Please select a location again.", "error")
        return redirect(url_for('location_input_page'))

    # Retrieve form data including new fields
    name = request.form.get('name')
    household_size = request.form.get('household_size')
    rooftop_area = request.form.get('rooftop_area')
    open_space_area = request.form.get('open_space_area')  # NEW
    roof_type = request.form.get('roof_type')
    property_type = request.form.get('property_type')  # NEW
    existing_water_sources = request.form.get('existing_water_sources')  # NEW
    intended_use = request.form.get('intended_use')  # NEW
    building_age = request.form.get('building_age')  # NEW
    occupancy = request.form.get('occupancy')  # NEW (optional – only for commercial)
    
    # --- Community Flow is now handled by /select_assessment ---
    # The old check 'if property_type == 'Community':' is no longer needed here.

    # Validate required fields
    name_valid, name_result = validate_name(name)
    if not name_valid:
        flash(name_result, "error")
        return redirect(url_for('individual_input_page'))
    
    # Validate household size
    size_valid, size_result = validate_household_size(household_size)
    if not size_valid:
        flash(size_result, "error")
        return redirect(url_for('individual_input_page'))
    
    # Validate rooftop area
    area_valid, area_result = validate_rooftop_area(rooftop_area)
    if not area_valid:
        flash(area_result, "error")
        return redirect(url_for('individual_input_page'))
    
    # Validate open space area
    space_valid, space_result = validate_open_space_area(open_space_area)
    if not space_valid:
        flash(space_result, "error")
        return redirect(url_for('individual_input_page'))
    
    # Convert validated values to appropriate types
    household_size = size_result
    rooftop_area = area_result
    open_space_area = space_result
    occupancy = int(occupancy) if occupancy else None
    
    # Create a new UserInput object with enhanced fields
    new_entry = UserInput(
        name=name,
        location_name=location_name,
        user_lat=user_lat,
        user_lon=user_lon,
        household_size=household_size,
        rooftop_area=rooftop_area,
        open_space_area=open_space_area,
        roof_type=roof_type,
        property_type=property_type,
        existing_water_sources=existing_water_sources,
        intended_use=intended_use,
        building_age=building_age,
        occupancy=occupancy
    )
    
    db.session.add(new_entry)
    db.session.commit()
    
    # Clear session data after successful submission
    session.pop('latitude', None)
    session.pop('longitude', None)
    session.pop('address', None)
    session.pop('assessment_type', None)

    # Reverting to a standard redirect, which works best with a native form submission
    # and is more reliable in avoiding browser navigation quirks.
    return redirect(url_for('results_page', entry_id=new_entry.id))

@app.route('/community-input')
def community_input_page():
    """Serves the community input page."""
    # Pass session data to the template if it exists
    return render_template('community-input.html',
                           location_name=session.get('address'),
                           user_lat=session.get('latitude'),
                           user_lon=session.get('longitude'))

@app.route('/submit_community_form', methods=['POST'])
def submit_community_form():
    # --- Retrieve location data from session ---
    user_lat = session.get('latitude')
    user_lon = session.get('longitude')
    location_name = session.get('address')
    property_type = session.get('assessment_type', 'Community') # Get from session

    if not all([user_lat, user_lon, location_name]):
        flash("Your session has expired. Please select a location again.", "error")
        return redirect(url_for('location_input_page'))

    # Basic info from hidden fields
    name = request.form.get('name', 'Community User')
    
    # Community-specific fields
    num_households = request.form.get('num_households', type=int)
    avg_household_size = request.form.get('avg_household_size', type=int)
    num_buildings = request.form.get('num_buildings', type=int)
    avg_rooftop_area = request.form.get('avg_rooftop_area', type=float)
    total_open_space = request.form.get('total_open_space', type=float)
    intended_use = request.form.get('intended_use')

    # --- Basic Validation ---
    if not all([location_name, user_lat, user_lon, num_households, num_buildings, avg_rooftop_area, total_open_space]):
        flash("Missing required community data.", "error")
        return redirect(url_for('community_input_page'))

    # Aggregate data for the main user input model
    total_household_size = num_households * avg_household_size
    total_rooftop_area = num_buildings * avg_rooftop_area

    # Create a new UserInput entry
    new_entry = UserInput(
        name=name,
        location_name=location_name,
        user_lat=float(user_lat),
        user_lon=float(user_lon),
        household_size=total_household_size,
        rooftop_area=total_rooftop_area,
        open_space_area=total_open_space,
        roof_type='mixed',  # Assume mixed for community
        property_type=property_type,
        existing_water_sources='Community Sources', # Placeholder
        intended_use=intended_use,
        building_age='mixed', # Placeholder
        occupancy=total_household_size # Use total population as occupancy
    )

    db.session.add(new_entry)
    db.session.commit()

    # Clear session data after successful submission
    session.pop('latitude', None)
    session.pop('longitude', None)
    session.pop('address', None)
    session.pop('assessment_type', None)

    return redirect(url_for('results_page', entry_id=new_entry.id))

@app.route('/results/<int:entry_id>')
def results_page(entry_id):
    # Redirect to the overview page of the results
    return redirect(url_for('results_overview', entry_id=entry_id))

@app.route('/results/property')
def property_details():
    entry_id = request.args.get('entry_id')
    user_data = UserInput.query.get_or_404(entry_id)
    
    try:
        location_analysis_data = get_api_data(user_data.user_lat, user_data.user_lon)
        if not location_analysis_data:
            flash("Unable to retrieve location data. Please check your coordinates and try again.", "error")
            return redirect(url_for('results_page', entry_id=entry_id))
    except Exception as e:
        print(f"Error fetching API data for property details: {e}")
        flash("An error occurred while analyzing your location. Please try again later.", "error")
        return redirect(url_for('results_page', entry_id=entry_id))
    
    comprehensive_analysis = calculate_comprehensive_feasibility(location_analysis_data, user_data)
    return render_template('property_details.html', user_data=user_data, location_data=location_analysis_data, analysis=comprehensive_analysis)

@app.route('/results/location')
def location_analysis():
    entry_id = request.args.get('entry_id')
    user_data = UserInput.query.get_or_404(entry_id)
    
    try:
        location_analysis_data = get_api_data(user_data.user_lat, user_data.user_lon)
        if not location_analysis_data:
            flash("Unable to retrieve location data. Please check your coordinates and try again.", "error")
            return redirect(url_for('results_page', entry_id=entry_id))
    except Exception as e:
        print(f"Error fetching API data for location analysis: {e}")
        flash("An error occurred while analyzing your location. Please try again later.", "error")
        return redirect(url_for('results_page', entry_id=entry_id))
    
    comprehensive_analysis = calculate_comprehensive_feasibility(location_analysis_data, user_data)
    return render_template('location_analysis.html', user_data=user_data, location_data=location_analysis_data, analysis=comprehensive_analysis)

@app.route('/results/hydrogeology')
def hydrogeological_profile():
    entry_id = request.args.get('entry_id')
    user_data = UserInput.query.get_or_404(entry_id)
    
    try:
        location_analysis_data = get_api_data(user_data.user_lat, user_data.user_lon)
        if not location_analysis_data:
            flash("Unable to retrieve location data. Please check your coordinates and try again.", "error")
            return redirect(url_for('results_page', entry_id=entry_id))
    except Exception as e:
        print(f"Error fetching API data for hydrogeological profile: {e}")
        flash("An error occurred while analyzing your location. Please try again later.", "error")
        return redirect(url_for('results_page', entry_id=entry_id))
    
    comprehensive_analysis = calculate_comprehensive_feasibility(location_analysis_data, user_data)
    return render_template('hydrogeological_profile.html', user_data=user_data, location_data=location_analysis_data, analysis=comprehensive_analysis)

@app.route('/results/feasibility')
def feasibility_assessment():
    entry_id = request.args.get('entry_id')
    user_data = UserInput.query.get_or_404(entry_id)
    
    try:
        location_analysis_data = get_api_data(user_data.user_lat, user_data.user_lon)
        if not location_analysis_data:
            flash("Unable to retrieve location data. Please check your coordinates and try again.", "error")
            return redirect(url_for('results_page', entry_id=entry_id))
    except Exception as e:
        print(f"Error fetching API data for feasibility assessment: {e}")
        flash("An error occurred while analyzing your location. Please try again later.", "error")
        return redirect(url_for('results_page', entry_id=entry_id))
    
    comprehensive_analysis = calculate_comprehensive_feasibility(location_analysis_data, user_data)
    return render_template('feasibility_assessment.html', user_data=user_data, location_data=location_analysis_data, analysis=comprehensive_analysis)

@app.route('/results/recommendations')
def recommendations():
    entry_id = request.args.get('entry_id')
    user_data = UserInput.query.get_or_404(entry_id)
    
    try:
        location_analysis_data = get_api_data(user_data.user_lat, user_data.user_lon)
        if not location_analysis_data:
            flash("Unable to retrieve location data. Please check your coordinates and try again.", "error")
            return redirect(url_for('results_page', entry_id=entry_id))
    except Exception as e:
        print(f"Error fetching API data for recommendations: {e}")
        flash("An error occurred while analyzing your location. Please try again later.", "error")
        return redirect(url_for('results_page', entry_id=entry_id))
    
    comprehensive_analysis = calculate_comprehensive_feasibility(location_analysis_data, user_data)
    return render_template('recommendations.html', user_data=user_data, location_data=location_analysis_data, analysis=comprehensive_analysis)

@app.route('/results/financials')
def financial_analysis():
    entry_id = request.args.get('entry_id')
    user_data = UserInput.query.get_or_404(entry_id)
    
    try:
        location_analysis_data = get_api_data(user_data.user_lat, user_data.user_lon)
        if not location_analysis_data:
            flash("Unable to retrieve location data. Please check your coordinates and try again.", "error")
            return redirect(url_for('results_page', entry_id=entry_id))
    except Exception as e:
        print(f"Error fetching API data for financial analysis: {e}")
        flash("An error occurred while analyzing your location. Please try again later.", "error")
        return redirect(url_for('results_page', entry_id=entry_id))
    
    comprehensive_analysis = calculate_comprehensive_feasibility(location_analysis_data, user_data)
    
    # Get detailed cost analysis using the category from comprehensive analysis
    cost_data = estimate_costs_and_payback(
        category_id=comprehensive_analysis['category']['category'],
        location_type=location_analysis_data.get('location_type', 'urban'),
        soil_type=location_analysis_data.get('soil_type', 'clay'),
        system_size=comprehensive_analysis['harvesting_potential']['annual_liters']
    )
    
    return render_template('financial_analysis.html', user_data=user_data, location_data=location_analysis_data, analysis=comprehensive_analysis, cost_data=cost_data)

## Removed legacy purification_page route

@app.route('/results/measurement_purification/<int:entry_id>')
def measurement_purification_page(entry_id):
    user_data = UserInput.query.get_or_404(entry_id)
    try:
        location_analysis_data = get_api_data(user_data.user_lat, user_data.user_lon)
        if not location_analysis_data:
            flash("Unable to retrieve location data. Please check your coordinates and try again.", "error")
            return redirect(url_for('results_page', entry_id=entry_id))
    except Exception as e:
        print(f"Error fetching API data for measurement & purification page: {e}")
        flash("An error occurred while analyzing your location. Please try again later.", "error")
        return redirect(url_for('results_page', entry_id=entry_id))

    comprehensive_analysis = calculate_comprehensive_feasibility(location_analysis_data, user_data)
    return render_template('measurement_purification.html', user_data=user_data, location_data=location_analysis_data, analysis=comprehensive_analysis)

@app.route('/results/awareness')
def awareness_page():
    return render_template('resources.html')

@app.route('/results/overview/<int:entry_id>')
def results_overview(entry_id):
    user_data = UserInput.query.get_or_404(entry_id)
    
    try:
        location_analysis_data = get_api_data(user_data.user_lat, user_data.user_lon)
        if not location_analysis_data:
            flash("Unable to retrieve location data. Please check your coordinates and try again.", "error")
            return redirect(url_for('results_page', entry_id=entry_id))
    except Exception as e:
        print(f"Error fetching API data for results overview: {e}")
        flash("An error occurred while analyzing your location. Please try again later.", "error")
        return redirect(url_for('results_page', entry_id=entry_id))
    
    comprehensive_analysis = calculate_comprehensive_feasibility(location_analysis_data, user_data)
    return render_template('results_overview.html', user_data=user_data, location_data=location_analysis_data, analysis=comprehensive_analysis)

@app.route('/download_report/<int:entry_id>')
def download_report(entry_id):
    # Retrieve user data and perform analysis (same logic as results_page)
    user_data = UserInput.query.get_or_404(entry_id)
    
    if not user_data.user_lat or not user_data.user_lon:
        return "Error: GPS coordinates are required for API-based analysis.", 400

    try:
        # Fetch combined data from all APIs using user's GPS
        location_data = get_api_data(user_data.user_lat, user_data.user_lon)
    except Exception as e:
        return f"Error during API data retrieval: {e}", 500

    if not location_data:
        return "Error: Could not find data for your location from APIs.", 404

    analysis = calculate_comprehensive_feasibility(location_data, user_data)

    # --- PDF Translation Setup ---
    lang = request.args.get('lang', 'en')
    translations_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'translations')
    translation_file_path = os.path.join(translations_dir, f'{lang}.json')

    if not os.path.exists(translation_file_path):
        # Fallback to English if the language file doesn't exist
        translation_file_path = os.path.join(translations_dir, 'en.json')

    with open(translation_file_path, 'r', encoding='utf-8') as f:
        t = json.load(f)

    # --- PDF Generation with FPDF2 - Now with Unicode font support ---
    class PDF(FPDF):
        def __init__(self, lang='en', *args, **kwargs):
            super().__init__(*args, **kwargs)
            font_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fonts')

            # --- Font setup based on language ---
            self.font_name = 'DejaVu' # Default font
            
            # Add DejaVu for English, Hindi, and as a fallback
            dejavu_sans_path = os.path.join(font_dir, 'DejaVuSans.ttf')
            dejavu_sans_bold_path = os.path.join(font_dir, 'DejaVuSans-Bold.ttf')
            self.add_font('DejaVu', '', dejavu_sans_path, uni=True)
            self.add_font('DejaVu', 'B', dejavu_sans_bold_path, uni=True)

            # If language is Telugu, add and switch to Noto Sans Telugu
            if lang == 'te':
                noto_telugu_path = os.path.join(font_dir, 'NotoSansTelugu-Regular.ttf')
                noto_telugu_bold_path = os.path.join(font_dir, 'NotoSansTelugu-Bold.ttf')
                if os.path.exists(noto_telugu_path) and os.path.exists(noto_telugu_bold_path):
                    self.add_font('NotoTelugu', '', noto_telugu_path, uni=True)
                    self.add_font('NotoTelugu', 'B', noto_telugu_bold_path, uni=True)
                    self.font_name = 'NotoTelugu' # Set as the active font
            
            self.set_font(self.font_name, '', 12)

        def header(self):
            self.set_font(self.font_name, 'B', 12)
            self.cell(0, 10, t['report_title'], new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            self.ln(5)

        def footer(self):
            self.set_y(-15)
            self.set_font(self.font_name, '', 8)
            self.cell(0, 10, f'Page {self.page_no()}', align='C')

        def section_title(self, title):
            self.set_font(self.font_name, 'B', 14)
            self.set_text_color(0, 77, 76)
            self.cell(0, 10, title, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
            self.line(self.get_x(), self.get_y(), self.get_x() + 190, self.get_y())
            self.ln(4)
        
        def write_key_value_table(self, data):
            self.set_font(self.font_name, '', 11)
            self.set_text_color(51, 51, 51)
            key_col_width = 65
            val_col_width = self.w - self.l_margin - self.r_margin - key_col_width
            line_height = self.font_size * 1.5
            for key, value in data.items():
                self.set_font(self.font_name, 'B')
                self.cell(key_col_width, line_height, key, border=0)
                self.set_font(self.font_name, '')
                self.multi_cell(val_col_width, line_height, str(value), border=0, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            self.ln(5)

        def write_list(self, items):
            self.set_font(self.font_name, '', 11)
            self.set_text_color(51, 51, 51)
            for item in items:
                self.multi_cell(0, 5, f'- {item}')
                self.ln(2)
            self.ln(5)

    pdf = PDF(lang=lang)
    pdf.add_page()
    pdf.set_font(pdf.font_name, 'B', 24)
    pdf.set_text_color(0, 77, 76)
    pdf.cell(0, 10, t['report_title'], new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font('DejaVu', '', 11)
    pdf.set_text_color(51, 51, 51)
    pdf.cell(0, 10, f"{t['generated_on'].split('{')[0]}{datetime.now().strftime('%d %B %Y')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(10)

    pdf.section_title(t['section1_title'])
    pdf.write_key_value_table({
        t['prop_owner']: user_data.name,
        t['location']: user_data.location_name,
        t['prop_type']: user_data.property_type,
        t['household_size']: f"{user_data.household_size} {t['people']}",
        t['rooftop_area']: f"{user_data.rooftop_area:.1f} m²",
        t['open_space_area']: f"{user_data.open_space_area:.1f} m²",
    })

    pdf.section_title(t['section2_title'])
    pdf.write_key_value_table({
        t['data_source']: t['data_for'].format(region=location_data['Region_Name']),
        t['annual_rainfall']: f"{location_data['Rainfall_mm']:.0f} mm",
        t['soil_type']: location_data['Soil_Type'],
        t['gw_depth']: f"{location_data['Groundwater_Depth_m']} {t['meters']}",
        t['dist_to_data']: f"{location_data['distance']:.1f} {t['km']}",
    })

    pdf.section_title(t['section3_title'])
    pdf.write_key_value_table({
        t['aquifer_type']: location_data['Aquifer_Type'],
        t['aquifer_depth']: f"{location_data['Aquifer_Depth_Min_m']} - {location_data['Aquifer_Depth_Max_m']} {t['meters']}",
        t['infiltration_rate']: f"{location_data['Infiltration_Rate_mm_per_hr']} mm/hr",
        t['water_quality']: location_data['Water_Quality'],
        t['remarks']: location_data['Remarks'],
    })

    pdf.section_title(t['section4_title'])
    pdf.write_key_value_table({
        t['annual_harvest_potential']: f"{analysis['harvesting_potential']['annual_liters']:,.0f} {t['liters']}",
        t['household_demand']: f"{analysis['annual_demand']:,.0f} {t['liters']}",
        t['demand_coverage']: f"{analysis['feasibility_percentage']}%",
        t['feasibility_status']: analysis['feasibility_status'],
    })

    pdf.section_title(t['section5_title'])
    pdf.write_key_value_table({
        t['category']: f"{t['category']} {analysis['category']['category']}: {analysis['category']['name']}",
        t['description']: analysis['category']['description'],
    })
    pdf.set_font('DejaVu', 'B', 11)
    pdf.cell(0, 10, t['rec_structures'], new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.write_list(analysis['category']['recommended_structures'])

    pdf.section_title(t['section6_title'])
    safety_status = t['safe'] if analysis['safety_check']['is_safe'] else t['caution']
    pdf.write_key_value_table({t['status']: safety_status})
    if not analysis['safety_check']['is_safe']:
        pdf.set_font('DejaVu', 'B', 11)
        pdf.cell(0, 10, t['potential_issues'], new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.write_list(analysis['safety_check']['safety_issues'])
        pdf.set_font('DejaVu', 'B', 11)
        pdf.cell(0, 10, t['alternatives'], new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.write_list(analysis['safety_check']['alternatives'])

    pdf.section_title(t['section7_title'])
    pdf.write_key_value_table({
        t['storage_tank_cap']: f"{analysis['structure_dimensions']['storage']['capacity_liters']:,.0f} {t['liters']}",
        t['storage_tank_cost']: analysis['structure_dimensions']['storage']['material_cost'],
    })
    if 'pit' in analysis['structure_dimensions']:
        pit = analysis['structure_dimensions']['pit']
        pdf.write_key_value_table({
            t['recharge_pit_dims']: f"{pit['length_m']}m x {pit['width_m']}m x {pit['depth_m']}m",
            t['recharge_pit_cost']: pit['material_cost'],
        })

    pdf.section_title(t['section8_title'])
    pdf.write_key_value_table({
        t['intended_use']: user_data.intended_use,
        t['maintenance_schedule']: analysis['purification']['maintenance_schedule'],
        t['est_treatment_cost']: analysis['purification']['estimated_cost'],
    })
    pdf.set_font('DejaVu', 'B', 11)
    pdf.cell(0, 10, t['rec_treatment_seq'], new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.write_list(analysis['purification']['treatment_sequence'])

    pdf.section_title(t['section9_title'])
    pdf.write_key_value_table({
        t['initial_investment']: f"₹{analysis['cost_analysis']['total_initial_cost']:,.0f}",
        t['gov_subsidy']: f"₹{analysis['cost_analysis']['subsidy_amount']:,.0f}",
        t['net_investment']: f"₹{analysis['cost_analysis']['net_investment']:,.0f}",
        t['annual_savings']: f"₹{analysis['cost_analysis']['annual_net_savings']:,.0f}",
        t['payback_period']: f"{analysis['cost_analysis']['payback_years']} {t['years']}",
        t['roi']: f"{analysis['cost_analysis']['roi_percentage']}%",
    })

    # The .output() method returns a bytearray, which we convert to bytes
    pdf_bytes = bytes(pdf.output())
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=RWH_Report_{user_data.name.replace(" ", "_")}.pdf'
    return response

@app.route('/api/calculate', methods=['POST'])
def api_calculate():
    """API endpoint for rapid calculations without database storage."""
    data = request.get_json()
    
    # Mock location data for API
    mock_location = {
        'Rainfall_mm': data.get('rainfall', 800),
        'Runoff_Coefficient': 0.8,
        'Groundwater_Depth_m': data.get('gw_depth', 10),
        'Soil_Type': data.get('soil_type', 'Loamy'),
        'Infiltration_Rate_mm_per_hr': data.get('infiltration', 15),
        'Water_Quality': data.get('water_quality', 'Good')
    }
    
    # Create mock user input
    class MockUser:
        def __init__(self, data):
            self.rooftop_area = data.get('roof_area', 100)
            self.open_space_area = data.get('open_space', 50)
            self.household_size = data.get('household_size', 4)
            self.roof_type = data.get('roof_type', 'Concrete')
            self.intended_use = data.get('intended_use', 'general')
    
    mock_user = MockUser(data)
    
    # Calculate comprehensive feasibility
    result = calculate_comprehensive_feasibility(mock_location, mock_user)
    
    return jsonify(result)

# --- ADMIN ROUTES ---

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if current_user.is_authenticated:
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = bool(request.form.get('remember'))
        
        if not username or not password:
            flash('Please enter both username and password.', 'error')
            return render_template('admin/login.html')
        
        user = AdminUser.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user, remember=remember)
            user.last_login = datetime.utcnow()
            db.session.commit()
            
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid username or password.', 'error')
    
    return render_template('admin/login.html')

@app.route('/admin/logout')
@login_required
def admin_logout():
    logout_user()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    # Get dashboard statistics
    total_users = UserInput.query.count()
    recent_users = UserInput.query.order_by(UserInput.id.desc()).limit(5).all()
    
    # Basic analytics
    stats = {
        'total_users': total_users,
        'total_admins': AdminUser.query.count(),
        'recent_signups': UserInput.query.filter(UserInput.id > max(0, total_users - 30)).count(),
        'popular_locations': db.session.query(UserInput.location_name, db.func.count(UserInput.location_name).label('count')).group_by(UserInput.location_name).order_by(db.text('count DESC')).limit(5).all()
    }
    
    return render_template('admin/dashboard.html', stats=stats, recent_users=recent_users)

@app.route('/admin/users')
@admin_required
def admin_users():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    
    query = UserInput.query
    if search:
        query = query.filter(UserInput.name.contains(search) | UserInput.location_name.contains(search))
    
    users = query.order_by(UserInput.id.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template('admin/users.html', users=users, search=search)

@app.route('/admin/users/<int:user_id>')
@admin_required
def admin_user_detail(user_id):
    user = UserInput.query.get_or_404(user_id)
    return render_template('admin/user_detail.html', user=user)

@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def admin_delete_user(user_id):
    user = UserInput.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    flash(f'User {user.name} has been deleted successfully.', 'success')
    return redirect(url_for('admin_users'))

def get_analytics_data():
    """
    Fetches and processes data for the analytics dashboard.
    This version performs aggregation in Python to avoid database-specific syntax.
    """
    # 1. Signups over time (by month)
    signups_by_month = {}
    all_signups = UserInput.query.with_entities(UserInput.created_at).all()
    for signup in all_signups:
        # Ensure created_at is a datetime object
        if isinstance(signup.created_at, datetime):
            month_key = signup.created_at.strftime('%Y-%m')
            signups_by_month[month_key] = signups_by_month.get(month_key, 0) + 1

    # Sort by key (date) to ensure chronological order
    sorted_signups = sorted(signups_by_month.items())
    signup_labels = [item[0] for item in sorted_signups]
    signup_values = [item[1] for item in sorted_signups]

    # 2. Property types distribution
    property_types_q = db.session.query(
        UserInput.property_type, 
        db.func.count(UserInput.property_type)
    ).group_by(UserInput.property_type).order_by(db.func.count(UserInput.property_type).desc()).all()
    property_labels = [item[0] if item[0] else 'Not Specified' for item in property_types_q]
    property_values = [item[1] for item in property_types_q]

    # 3. Average rooftop area by property type
    avg_rooftop_q = db.session.query(
        UserInput.property_type,
        db.func.avg(UserInput.rooftop_area)
    ).group_by(UserInput.property_type).all()
    avg_rooftop_labels = [item[0] if item[0] else 'Not Specified' for item in avg_rooftop_q]
    avg_rooftop_values = [round(item[1], 2) if item[1] else 0 for item in avg_rooftop_q]

    # 4. Geographic distribution (top 10 cities)
    top_cities_q = db.session.query(
        UserInput.location_name,
        db.func.count(UserInput.location_name)
    ).group_by(UserInput.location_name).order_by(db.func.count(UserInput.location_name).desc()).limit(10).all()
    city_labels = [item[0] for item in top_cities_q]
    city_values = [item[1] for item in top_cities_q]

    return {
        'signups_over_time': {
            'labels': signup_labels,
            'data': signup_values
        },
        'property_type_distribution': {
            'labels': property_labels,
            'data': property_values
        },
        'avg_rooftop_area': {
            'labels': avg_rooftop_labels,
            'data': avg_rooftop_values
        },
        'geographic_distribution': {
            'labels': city_labels,
            'data': city_values
        }
    }

@app.route('/admin/analytics')
@admin_required
def admin_analytics():
    analytics_data = get_analytics_data()
    return render_template('admin/analytics.html', analytics=analytics_data)

@app.route('/interactive-map')
@app.route('/interactive-map/<int:entry_id>')
def interactive_map_page(entry_id=None):
    """Serves the interactive map page, focusing on a single location if entry_id is provided."""
    user_location_details = None
    nearest_location_data = None

    if entry_id:
        user_data = UserInput.query.get(entry_id)
        if user_data:
            # Find the single nearest data point from the CSV
            if user_data.user_lat and user_data.user_lon:
                nearest_location_data = get_nearest_location(user_data.user_lat, user_data.user_lon)
            else: # Fallback to manual location name if no GPS
                nearest_location_data = get_mock_location_data(user_data.location_name)

            # Create a new object for the user's location that includes the derived environmental data
            if nearest_location_data and user_data.user_lat and user_data.user_lon:
                user_location_details = nearest_location_data.copy() # Start with all data from nearest point
                user_location_details['Latitude'] = user_data.user_lat
                user_location_details['Longitude'] = user_data.user_lon
                user_location_details['Region_Name'] = "Your Location" # Label it clearly

    # Pass both the user's detailed location and the nearest CSV point
    return render_template('interactive_map.html', user_location=user_location_details, nearest_location=nearest_location_data)


@app.route('/admin/export/users')
@admin_required
def admin_export_users():
    """Export all user data as CSV"""
    import csv
    import io
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['ID', 'Name', 'Location', 'Latitude', 'Longitude', 'Household Size', 
                     'Rooftop Area', 'Open Space Area', 'Roof Type', 'Property Type', 
                     'Intended Use'])
    
    # Write data
    users = UserInput.query.all()
    for user in users:
        writer.writerow([
            user.id, user.name, user.location_name, user.user_lat, user.user_lon,
            user.household_size, user.rooftop_area, user.open_space_area, 
            user.roof_type, user.property_type, user.intended_use
        ])
    
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    
    return response

# --- API Configuration ---
# IMPORTANT: Replace with your actual API key from OpenWeatherMap
OPENWEATHERMAP_API_KEY = os.environ.get('OPENWEATHERMAP_API_KEY', '32dc29dca01bde623300f501d45e42dd')
SOILGRIDS_API_ENDPOINT = "https://rest.isric.org/soilgrids/v2.0/properties/query"

def get_monthly_rainfall_data(lat, lon, api_key):
    """
    Fetches monthly rainfall breakdown from OpenWeatherMap 5-day forecast.
    Returns a dictionary with monthly estimates.
    """
    try:
        forecast_url = f"https://api.openweathermap.org/data/2.5/forecast?lat={lat}&lon={lon}&appid={api_key}&units=metric"
        response = requests.get(forecast_url, timeout=10)
        response.raise_for_status()
        data = response.json()

        # Group rainfall by month
        monthly_rain = {}
        for item in data.get('list', []):
            dt = item.get('dt_txt', '')
            if dt:
                month = dt.split('-')[1]  # Extract month from YYYY-MM-DD HH:MM:SS
                rain = item.get('rain', {}).get('3h', 0)
                monthly_rain[month] = monthly_rain.get(month, 0) + rain

        # Convert 3-hour data to monthly estimates (rough approximation)
        # Since we only have 5 days, we'll use regional patterns to estimate full months
        regional_avg = get_location_specific_rainfall_fallback(lat, lon)

        # Estimate monthly distribution based on regional patterns
        monthly_distribution = {
            '01': 0.08, '02': 0.07, '03': 0.07, '04': 0.06, '05': 0.08, '06': 0.10,  # Dry season
            '07': 0.15, '08': 0.15, '09': 0.12, '10': 0.08, '11': 0.03, '12': 0.01   # Monsoon season
        }

        monthly_breakdown = {}
        for month, fraction in monthly_distribution.items():
            monthly_breakdown[month] = regional_avg * fraction

        return monthly_breakdown

    except Exception as e:
        print(f"Error fetching monthly rainfall data: {e}")
        # Return default monthly distribution
        regional_avg = get_location_specific_rainfall_fallback(lat, lon)
        return {month: regional_avg * frac for month, frac in {
            '01': 0.08, '02': 0.07, '03': 0.07, '04': 0.06, '05': 0.08, '06': 0.10,
            '07': 0.15, '08': 0.15, '09': 0.12, '10': 0.08, '11': 0.03, '12': 0.01
        }.items()}

def get_live_weather_data(lat, lon, api_key):
    """
    Fetches current weather conditions including temperature, humidity, and current rain.
    """
    try:
        current_url = f"https://api.openweathermap.org/data/2.5/weather?lat={lat}&lon={lon}&appid={api_key}&units=metric"
        response = requests.get(current_url, timeout=10)
        response.raise_for_status()
        data = response.json()

        return {
            'temperature': data.get('main', {}).get('temp'),
            'humidity': data.get('main', {}).get('humidity'),
            'pressure': data.get('main', {}).get('pressure'),
            'current_rain': data.get('rain', {}).get('1h', 0) or data.get('rain', {}).get('3h', 0),
            'weather_description': data.get('weather', [{}])[0].get('description', ''),
            'wind_speed': data.get('wind', {}).get('speed'),
            'location_name': data.get('name', 'Unknown')
        }

    except Exception as e:
        print(f"Error fetching live weather data: {e}")
        return {
            'temperature': 25,
            'humidity': 60,
            'pressure': 1013,
            'current_rain': 0,
            'weather_description': 'Data unavailable',
            'wind_speed': 5,
            'location_name': 'Unknown'
        }
    """
    Returns location-specific rainfall fallback values based on regional climate patterns in India.
    Uses latitude/longitude to determine the region and return appropriate average rainfall.
    """
    # Define regional rainfall patterns (approximate annual averages in mm)
    # Adjusted boundaries to better cover Indian geography
    regional_rainfall = {
        # North India (Delhi, Punjab, Haryana, UP, Rajasthan border areas)
        'north': {'lat_range': (26, 35), 'lon_range': (70, 85), 'rainfall': 700},
        # North-East India (Assam, Meghalaya, etc.)
        'northeast': {'lat_range': (24, 29), 'lon_range': (85, 98), 'rainfall': 2500},
        # East India (West Bengal, Odisha, Bihar)
        'east': {'lat_range': (20, 27), 'lon_range': (82, 90), 'rainfall': 1500},
        # Central India (Madhya Pradesh, Chhattisgarh)
        'central': {'lat_range': (18, 26), 'lon_range': (75, 85), 'rainfall': 1000},
        # West India (Maharashtra, Gujarat, Rajasthan west)
        'west': {'lat_range': (15, 26), 'lon_range': (68, 76), 'rainfall': 1000},
        # South India (Kerala, Karnataka, Tamil Nadu, Andhra)
        'south': {'lat_range': (8, 18), 'lon_range': (70, 85), 'rainfall': 2000},
        # North-West (Rajasthan dry regions)
        'northwest': {'lat_range': (24, 30), 'lon_range': (70, 75), 'rainfall': 300},
    }

    # Determine region based on coordinates
    for region, data in regional_rainfall.items():
        lat_min, lat_max = data['lat_range']
        lon_min, lon_max = data['lon_range']

        if lat_min <= lat <= lat_max and lon_min <= lon <= lon_max:
            rainfall = data['rainfall']
            print(f"Location ({lat:.2f}, {lon:.2f}) falls in {region} region - using {rainfall}mm rainfall fallback")
            return rainfall

    # Default fallback if coordinates don't match any region
    print(f"Location ({lat:.2f}, {lon:.2f}) not matched to specific region - using 1000mm default")
    return 1000

def get_location_specific_rainfall_fallback(lat, lon):
    """
    Returns location-specific rainfall fallback values based on regional climate patterns in India.
    Uses latitude/longitude to determine the region and return appropriate average rainfall.
    """
    # Define regional rainfall patterns (approximate annual averages in mm)
    # Adjusted boundaries to better cover Indian geography
    regional_rainfall = {
        # North India (Delhi, Punjab, Haryana, UP, Rajasthan border areas)
        'north': {'lat_range': (26, 35), 'lon_range': (70, 85), 'rainfall': 700},
        # North-East India (Assam, Meghalaya, etc.)
        'northeast': {'lat_range': (24, 29), 'lon_range': (85, 98), 'rainfall': 2500},
        # East India (West Bengal, Odisha, Bihar)
        'east': {'lat_range': (20, 27), 'lon_range': (82, 90), 'rainfall': 1500},
        # Central India (Madhya Pradesh, Chhattisgarh)
        'central': {'lat_range': (18, 26), 'lon_range': (75, 85), 'rainfall': 1000},
        # West India (Maharashtra, Gujarat, Rajasthan west)
        'west': {'lat_range': (15, 26), 'lon_range': (68, 76), 'rainfall': 1000},
        # South India (Kerala, Karnataka, Tamil Nadu, Andhra)
        'south': {'lat_range': (8, 18), 'lon_range': (70, 85), 'rainfall': 2000},
        # North-West (Rajasthan dry regions)
        'northwest': {'lat_range': (24, 30), 'lon_range': (70, 75), 'rainfall': 300},
    }

    # Determine region based on coordinates
    for region, data in regional_rainfall.items():
        lat_min, lat_max = data['lat_range']
        lon_min, lon_max = data['lon_range']

        if lat_min <= lat <= lat_max and lon_min <= lon <= lon_max:
            rainfall = data['rainfall']
            print(f"Location ({lat:.2f}, {lon:.2f}) falls in {region} region - using {rainfall}mm rainfall fallback")
            return rainfall

    # Default fallback if coordinates don't match any region
    print(f"Location ({lat:.2f}, {lon:.2f}) not matched to specific region - using 1000mm default")
    return 1000

def get_rainfall_from_api(lat, lon, api_key):
    """
    Fetches rainfall data from OpenWeatherMap APIs.
    Uses multiple approaches: 5-day forecast, current weather, and regional fallbacks.
    """
    # First, try to get rainfall from 5-day forecast (free tier)
    try:
        forecast_url = f"https://api.openweathermap.org/data/2.5/forecast?lat={lat}&lon={lon}&appid={api_key}&units=metric"
        response = requests.get(forecast_url, timeout=10)
        response.raise_for_status()
        data = response.json()

        # Extract rainfall from forecast (3-hourly data for 5 days = 40 entries)
        total_rainfall = 0
        rain_days = 0

        for item in data.get('list', []):
            rain = item.get('rain', {}).get('3h', 0)  # rainfall in last 3 hours
            if rain > 0:
                total_rainfall += rain
                rain_days += 1

        # If we have rainfall data, extrapolate to annual estimate
        if total_rainfall > 0:
            # 5 days of data, extrapolate to annual (365 days)
            # But adjust for the fact that not all days rain
            annual_estimate = (total_rainfall / 5) * 365 * (rain_days / len(data.get('list', [])))
            print(f"Forecast-based rainfall estimate: {annual_estimate:.1f}mm/year")
            return max(annual_estimate, 100)  # Minimum 100mm to avoid unrealistic values

    except Exception as e:
        print(f"Forecast API failed: {e}")

    # Fallback: Try current weather for any rain information
    try:
        current_url = f"https://api.openweathermap.org/data/2.5/weather?lat={lat}&lon={lon}&appid={api_key}&units=metric"
        response = requests.get(current_url, timeout=10)
        response.raise_for_status()
        data = response.json()

        # Check for current rain
        rain = data.get('rain', {}).get('1h', 0) or data.get('rain', {}).get('3h', 0)
        if rain > 0:
            print(f"Current rain detected: {rain}mm - using regional average")
            # If it's currently raining, use regional average
            return get_location_specific_rainfall_fallback(lat, lon)

    except Exception as e:
        print(f"Current weather API failed: {e}")

    # Final fallback: Use regional averages
    return get_location_specific_rainfall_fallback(lat, lon)

def get_soil_data_from_api(lat, lon):
    """
    Fetches soil properties from the ISRIC SoilGrids API.
    """
    params = {
        "lat": lat,
        "lon": lon,
        "properties": "clay,sand,silt", # Request clay, sand, and silt content
        "depths": "0-5cm", # Focus on topsoil
    }
    try:
        response = requests.get(SOILGRIDS_API_ENDPOINT, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        
        properties = data.get('properties', {}).get('layers', [{}])[0].get('depths', [{}])[0].get('values', {})
        
        # Extract mean values, with fallbacks
        clay_content = properties.get('clay', {}).get('mean', 250) / 10  # Convert from g/kg to %
        sand_content = properties.get('sand', {}).get('mean', 400) / 10
        silt_content = properties.get('silt', {}).get('mean', 350) / 10

        # Determine soil type based on composition
        if sand_content > 50:
            soil_type = "Sandy"
            infiltration_rate = 20
        elif clay_content > 40:
            soil_type = "Clayey"
            infiltration_rate = 5
        else:
            soil_type = "Loamy"
            infiltration_rate = 15

        return {
            "Soil_Type": soil_type,
            "Infiltration_Rate_mm_per_hr": infiltration_rate,
            "Soil_Permability_Class": "High" if infiltration_rate > 15 else "Medium"
        }
    except requests.exceptions.Timeout:
        print(f"Timeout error fetching soil data from ISRIC for location ({lat}, {lon})")
        return {
            "Soil_Type": "Loamy",
            "Infiltration_Rate_mm_per_hr": 15,
            "Soil_Permability_Class": "Medium"
        }
    except requests.exceptions.ConnectionError:
        print(f"Connection error fetching soil data from ISRIC for location ({lat}, {lon})")
        return {
            "Soil_Type": "Loamy",
            "Infiltration_Rate_mm_per_hr": 15,
            "Soil_Permability_Class": "Medium"
        }
    except requests.exceptions.HTTPError as e:
        print(f"HTTP error fetching soil data from ISRIC: {e}")
        return {
            "Soil_Type": "Loamy",
            "Infiltration_Rate_mm_per_hr": 15,
            "Soil_Permability_Class": "Medium"
        }
    except Exception as e:
        print(f"Unexpected error fetching soil data from ISRIC: {e}")
        return {

            "Soil_Type": "Loamy",
            "Infiltration_Rate_mm_per_hr": 15,
            "Soil_Permability_Class": "Medium"
        }

def get_api_data(lat, lon):
    """
    Main function to fetch and combine data from all external APIs.
    Checks database first, fetches from APIs if not found, then stores for future use.
    """
    # First, check if we already have data for this location in the database
    existing_data = GeoData.query.filter_by(latitude=lat, longitude=lon).first()
    if existing_data:
        print(f"Using cached API data for location ({lat}, {lon})")
        # Convert database record to the expected format
        return {
            "Rainfall_mm": existing_data.rainfall_mm,
            "Monthly_Rainfall_mm": {},  # Not cached, will be fetched fresh
            "Live_Weather": {},  # Not cached, will be fetched fresh
            "Soil_Type": existing_data.soil_type,
            "Infiltration_Rate_mm_per_hr": existing_data.infiltration_rate_mm_per_hr,
            "Soil_Permability_Class": existing_data.soil_permability_class,
            "Groundwater_Depth_m": existing_data.groundwater_depth_m,
            "Aquifer_Type": existing_data.aquifer_type,
            "Aquifer_Depth_Min_m": existing_data.aquifer_depth_min_m,
            "Aquifer_Depth_Max_m": existing_data.aquifer_depth_max_m,
            "Aquifer_Thickness_m": existing_data.aquifer_thickness_m,
            "Water_Quality": existing_data.water_quality,
            "Remarks": existing_data.remarks,
            "Region_Name": existing_data.region_name,
            "distance": 0,  # Not stored in DB
            "Runoff_Coefficient": 0.85,
            "Water_Cost_per_Liter": existing_data.water_cost_per_liter,
            "Aquifer_Material_State": "Unknown",  # Would need additional query
            "Aquifer_Material_Type": "Unknown",
            "Aquifer_Material_Area": 0,
        }

    print(f"Fetching fresh API data for location ({lat}, {lon})")

    # 1. Get live rainfall data
    rainfall = get_rainfall_from_api(lat, lon, OPENWEATHERMAP_API_KEY)

    # 2. Get monthly rainfall breakdown
    monthly_rainfall = get_monthly_rainfall_data(lat, lon, OPENWEATHERMAP_API_KEY)

    # 3. Get live weather conditions
    live_weather = get_live_weather_data(lat, lon, OPENWEATHERMAP_API_KEY)

    # 4. Get live soil data
    soil_data = get_soil_data_from_api(lat, lon)

    # 3. Get fallback data from the database for groundwater, aquifer, etc.
    fallback_data = get_nearest_geo_data_from_db(lat, lon)
    if not fallback_data:
        # If DB lookup fails, create a default fallback structure
        fallback_data = {
            'groundwater_depth_m': 10, 'aquifer_type': 'Unconfined', 'aquifer_depth_min_m': 10,
            'aquifer_depth_max_m': 30, 'aquifer_thickness_m': 20, 'water_quality': 'Good',
            'remarks': 'API data - no nearby station data', 'region_name': f'Location ({lat:.4f}, {lon:.4f})', 'distance': 0,
            'water_cost_per_liter': 0.16
        }

    # 4. Get aquifer material data from PostGIS
    aquifer_data = get_aquifer_material_at_location(lat, lon)

    # 5. Combine all data into a single dictionary
    combined_data = {
        "Rainfall_mm": rainfall,
        "Monthly_Rainfall_mm": monthly_rainfall,
        "Live_Weather": live_weather,
        "Soil_Type": soil_data["Soil_Type"],
        "Infiltration_Rate_mm_per_hr": soil_data["Infiltration_Rate_mm_per_hr"],
        "Soil_Permability_Class": soil_data["Soil_Permability_Class"],
        # --- Fields from database fallback ---
        "Groundwater_Depth_m": fallback_data.get('groundwater_depth_m', 10),
        "Aquifer_Type": fallback_data.get('aquifer_type', 'Unconfined'),
        "Aquifer_Depth_Min_m": fallback_data.get('aquifer_depth_min_m', 10),
        "Aquifer_Depth_Max_m": fallback_data.get('aquifer_depth_max_m', 30),
        "Aquifer_Thickness_m": fallback_data.get('aquifer_thickness_m', 20),
        "Water_Quality": fallback_data.get('water_quality', 'Good'),
        "Remarks": fallback_data.get('remarks', 'Data from APIs and nearest station'),
        "Region_Name": fallback_data.get('region_name', f'Location ({lat:.4f}, {lon:.4f})'),
        "distance": fallback_data.get('distance', 0),
        "Runoff_Coefficient": 0.85, # Default, can be adjusted based on roof type later
        "Water_Cost_per_Liter": fallback_data.get('water_cost_per_liter', 0.16),
        # --- New PostGIS aquifer data ---
        "Aquifer_Material_State": aquifer_data.get('state_name', 'Unknown') if aquifer_data.get('found') else 'Not Available',
        "Aquifer_Material_Type": aquifer_data.get('aquifer_type', 'Unknown') if aquifer_data.get('found') else 'Not Available',
        "Aquifer_Material_Area": aquifer_data.get('area', 0) if aquifer_data.get('found') else 0,
    }

    # 6. Save the API data to database for future use
    try:
        geo_entry = GeoData(
            region_name=combined_data["Region_Name"],
            state="Unknown",  # We don't get state from APIs
            latitude=lat,
            longitude=lon,
            rainfall_mm=combined_data["Rainfall_mm"],
            groundwater_depth_m=combined_data["Groundwater_Depth_m"],
            aquifer_type=combined_data["Aquifer_Type"],
            aquifer_depth_min_m=combined_data["Aquifer_Depth_Min_m"],
            aquifer_depth_max_m=combined_data["Aquifer_Depth_Max_m"],
            aquifer_thickness_m=combined_data["Aquifer_Thickness_m"],
            remarks=combined_data["Remarks"],
            soil_type=combined_data["Soil_Type"],
            infiltration_rate_mm_per_hr=combined_data["Infiltration_Rate_mm_per_hr"],
            soil_permability_class=combined_data["Soil_Permability_Class"],
            water_quality=combined_data["Water_Quality"],
            water_cost_per_liter=combined_data["Water_Cost_per_Liter"]
        )
        db.session.add(geo_entry)
        db.session.commit()
        print(f"Saved API data for location ({lat}, {lon}) to database")
    except Exception as e:
        print(f"Error saving API data to database: {e}")
        db.session.rollback()

    return combined_data

# def load_csv_to_db():
#     """
#     Loads data from the mock_location_data.csv file into the GeoData table.
#     This function should be run once to populate the database.
#     It checks if data already exists to prevent duplicate entries.
#     """
#     with app.app_context():
#         if GeoData.query.first() is not None:
#             print("Geological data already exists in the database. Skipping import.")
#             return

#         print("Importing geological data from CSV to database...")
#         try:
#             df = pd.read_csv(CSV_FILE_PATH)
#             # Rename columns to match the GeoData model exactly
#             df.rename(columns={
#                 'Region_Name': 'region_name',
#                 'State': 'state',
#                 'Latitude': 'latitude',
#                 'Longitude': 'longitude',
#                 'Rainfall_mm': 'rainfall_mm',
#                 'Groundwater_Depth_m': 'groundwater_depth_m',
#                 'Aquifer_Type': 'aquifer_type',
#                 'Aquifer_Depth_Min_m': 'aquifer_depth_min_m',
#                 'Aquifer_Depth_Max_m': 'aquifer_depth_max_m',
#                 'Aquifer_Thickness_m': 'aquifer_thickness_m',
#                 'Remarks': 'remarks',
#                 'Soil_Type': 'soil_type',
#                 'Infiltration_Rate_mm_per_hr': 'infiltration_rate_mm_per_hr',
#                 'Soil_Permability_Class': 'soil_permability_class',
#                 'Water_Quality': 'water_quality',
#                 'Water_Cost_per_Liter': 'water_cost_per_liter'
#             }, inplace=True)

#             # Ensure all columns exist, fill missing with None (which becomes NULL in DB)
#             for col in GeoData.__table__.columns.keys():
#                 if col not in df.columns and col != 'id':
#                     df[col] = None

#             # Convert to records and add to database
#             records = df.to_dict(orient='records')
#             for record in records:
#                 # Filter out any keys that are not in the model
#                 filtered_record = {k: v for k, v in record.items() if k in GeoData.__table__.columns.keys()}
#                 geo_entry = GeoData(**filtered_record)
#                 db.session.add(geo_entry)
            
#             db.session.commit()
#             print(f"Successfully imported {len(records)} records into the GeoData table.")
#         except FileNotFoundError:
#             print(f"CRITICAL ERROR: Could not find '{CSV_FILE_PATH}' to populate the database.")
#         except Exception as e:
#             print(f"An error occurred during CSV import: {e}")
#             db.session.rollback()

def get_nearest_geo_data_from_db(lat, lon):
    """
    Finds the nearest geological data point from the GeoData table in the database.
    """
    # This is a simplified distance calculation. For production, consider using PostGIS for efficiency.
    all_geo_data = GeoData.query.all()
    if not all_geo_data:
        return None

    min_dist = float('inf')
    nearest_data = None

    for location in all_geo_data:
        dist = haversine(lat, lon, location.latitude, location.longitude)
        if dist < min_dist:
            min_dist = dist
            nearest_data = location

    if nearest_data:
        result = nearest_data.to_dict()
        result['distance'] = min_dist
        return result
    return None

def get_aquifer_material_at_location(lat, lon):
    """
    Query the PostGIS database to find aquifer material data at a specific latitude and longitude.
    Returns the aquifer material information if found, None otherwise.
    """
    try:
        # Create a point geometry from the lat/lon coordinates
        point = WKTElement(f'POINT({lon} {lat})', srid=4326)
        
        # Query for aquifer material that contains this point
        result = db.session.query(AquiferMaterial).filter(
            func.ST_Contains(AquiferMaterial.geometry, point)
        ).first()
        
        if result:
            return {
                'state_name': result.name_of_st,
                'aquifer_type': result.type_of_aq,
                'area': result.st_area_sh,
                'length': result.st_length_,
                'found': True
            }
        else:
            return {'found': False, 'message': 'No aquifer material data found at this location'}
            
    except Exception as e:
        print(f"Error querying aquifer material data: {e}")
        return {'found': False, 'message': f'Database error: {str(e)}'}

if __name__ == '__main__':
    with app.app_context():
        # Create the database tables if they don't exist
        db.create_all()
        
        # Load data from CSV into the database
        # load_csv_to_db()  # Commented out - CSV file not available
        
        # Create default admin user if no admin exists
        if not AdminUser.query.first():
            admin = AdminUser(
                username='admin',
                email='admin@rainwaterharvesting.com',
                role='admin'
            )
            admin.set_password('admin123')  # Change this password!
            db.session.add(admin)
            db.session.commit()
            print("Default admin user created:")
            print("Username: admin")
            print("Password: admin123")
            print("Please change this password after first login!")
    
    app.run(debug=True, host='0.0.0.0', port=5000)

# API Routes for Interactive Map
@app.route('/api/groundwater-stations')
def get_groundwater_stations():
    """API endpoint to get ground water station data for the map."""
    try:
        from models import GroundWaterLevelStation
        stations = GroundWaterLevelStation.query.limit(1000).all()  # Limit for performance
        
        data = []
        for station in stations:
            if station.lat and station.long:
                data.append({
                    'id': station.id,
                    'name': station.station_na or 'Unknown',
                    'lat': station.lat,
                    'lng': station.long,
                    'state': station.state_name,
                    'district': station.district_n,
                    'agency': station.agency_nam,
                    'basin': station.basin_name
                })
        
        return jsonify({'stations': data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/aquifers')
def get_aquifers():
    """API endpoint to get major aquifer data for the map."""
    try:
        from models import MajorAquifer
        aquifers = MajorAquifer.query.limit(500).all()  # Limit for performance
        
        data = []
        for aquifer in aquifers:
            # For polygon geometries, we'll need to extract coordinates
            # For now, just return basic info
            data.append({
                'id': aquifer.id,
                'name': aquifer.aquifer or 'Unknown',
                'state': aquifer.state,
                'system': aquifer.system,
                'zone': aquifer.zone_m
            })
        
        return jsonify({'aquifers': data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/geo/groundwater')
def api_geo_groundwater():
    """GeoJSON endpoint for Ground Water Level Stations, optionally filtered by lat/lon and radius_km."""
    try:
        from sqlalchemy import func
        import json as _json
        from models import GroundWaterLevelStation

        lat = request.args.get('lat', type=float)
        lon = request.args.get('lon', type=float)
        radius_km = request.args.get('radius_km', default=200, type=float)

        query = db.session.query(
            GroundWaterLevelStation.id,
            GroundWaterLevelStation.station_na,
            GroundWaterLevelStation.state_name,
            GroundWaterLevelStation.district_n,
            GroundWaterLevelStation.agency_nam,
            GroundWaterLevelStation.basin_name,
            GroundWaterLevelStation.geometry
        )

        if lat is not None and lon is not None:
            point = func.ST_SetSRID(func.ST_MakePoint(lon, lat), 4326)
            query = query.filter(
                func.ST_DWithin(
                    GroundWaterLevelStation.geometry.cast(db.text('geography')),
                    point.cast(db.text('geography')),
                    radius_km * 1000
                )
            )

        rows = query.limit(2000).all()
        features = []
        for r in rows:
            geom_json = db.session.scalar(db.func.ST_AsGeoJSON(r.geometry))
            if not geom_json:
                continue
            features.append({
                'type': 'Feature',
                'geometry': _json.loads(geom_json),
                'properties': {
                    'id': r.id,
                    'name': r.station_na or 'Unknown',
                    'state': r.state_name,
                    'district': r.district_n,
                    'agency': r.agency_nam,
                    'basin': r.basin_name
                }
            })

        return jsonify({'type': 'FeatureCollection', 'features': features})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/geo/aquifers')
def api_geo_aquifers():
    """GeoJSON endpoint for Major Aquifers polygons, optionally filtered by lat/lon and radius_km."""
    try:
        from sqlalchemy import func
        import json as _json
        from models import MajorAquifer

        lat = request.args.get('lat', type=float)
        lon = request.args.get('lon', type=float)
        radius_km = request.args.get('radius_km', default=250, type=float)

        query = db.session.query(
            MajorAquifer.id,
            MajorAquifer.aquifer,
            MajorAquifer.state,
            MajorAquifer.system,
            MajorAquifer.zone_m,
            MajorAquifer.geometry
        )

        if lat is not None and lon is not None:
            point = func.ST_SetSRID(func.ST_MakePoint(lon, lat), 4326)
            buffer = func.ST_Buffer(point.cast(db.text('geography')), radius_km * 1000)
            query = query.filter(func.ST_Intersects(MajorAquifer.geometry.cast(db.text('geography')), buffer))

        rows = query.limit(1500).all()
        features = []
        for r in rows:
            geom_json = db.session.scalar(db.func.ST_AsGeoJSON(r.geometry))
            if not geom_json:
                continue
            features.append({
                'type': 'Feature',
                'geometry': _json.loads(geom_json),
                'properties': {
                    'id': r.id,
                    'name': r.aquifer or 'Unknown',
                    'state': r.state,
                    'system': r.system,
                    'zone': r.zone_m
                }
            })

        return jsonify({'type': 'FeatureCollection', 'features': features})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/geo/aquifer-materials')
def api_geo_aquifer_materials():
    """GeoJSON endpoint for Aquifer Materials polygons, optionally filtered by lat/lon and radius_km."""
    try:
        from sqlalchemy import func
        import json as _json
        from models import AquiferMaterial

        lat = request.args.get('lat', type=float)
        lon = request.args.get('lon', type=float)
        radius_km = request.args.get('radius_km', default=250, type=float)

        query = db.session.query(
            AquiferMaterial.id,
            AquiferMaterial.Name_of_St,
            AquiferMaterial.Type_of_Aq,
            AquiferMaterial.st_area_sh,
            AquiferMaterial.st_length_,
            AquiferMaterial.geometry
        )

        if lat is not None and lon is not None:
            point = func.ST_SetSRID(func.ST_MakePoint(lon, lat), 4326)
            buffer = func.ST_Buffer(point.cast(db.text('geography')), radius_km * 1000)
            query = query.filter(func.ST_Intersects(AquiferMaterial.geometry.cast(db.text('geography')), buffer))

        rows = query.limit(1500).all()
        features = []
        for r in rows:
            geom_json = db.session.scalar(db.func.ST_AsGeoJSON(r.geometry))
            if not geom_json:
                continue
            features.append({
                'type': 'Feature',
                'geometry': _json.loads(geom_json),
                'properties': {
                    'id': r.id,
                    'state': r.Name_of_St,
                    'material_type': r.Type_of_Aq,
                    'area': r.st_area_sh,
                    'length': r.st_length_
                }
            })

        return jsonify({'type': 'FeatureCollection', 'features': features})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/geo/gw-quality')
def api_geo_gw_quality():
    """GeoJSON endpoint for Ground Water Quality Stations, optionally filtered by lat/lon and radius_km."""
    try:
        from sqlalchemy import func
        import json as _json
        from models import GroundWaterQualityStation

        lat = request.args.get('lat', type=float)
        lon = request.args.get('lon', type=float)
        radius_km = request.args.get('radius_km', default=200, type=float)

        query = db.session.query(
            GroundWaterQualityStation.id,
            GroundWaterQualityStation.station_na,
            GroundWaterQualityStation.state_name,
            GroundWaterQualityStation.district_n,
            GroundWaterQualityStation.agency_nam,
            GroundWaterQualityStation.basin_name,
            GroundWaterQualityStation.geometry
        )

        if lat is not None and lon is not None:
            point = func.ST_SetSRID(func.ST_MakePoint(lon, lat), 4326)
            query = query.filter(
                func.ST_DWithin(
                    GroundWaterQualityStation.geometry.cast(db.text('geography')),
                    point.cast(db.text('geography')),
                    radius_km * 1000
                )
            )

        rows = query.limit(2000).all()
        features = []
        for r in rows:
            geom_json = db.session.scalar(db.func.ST_AsGeoJSON(r.geometry))
            if not geom_json:
                continue
            features.append({
                'type': 'Feature',
                'geometry': _json.loads(geom_json),
                'properties': {
                    'id': r.id,
                    'name': r.station_na or 'Unknown',
                    'state': r.state_name,
                    'district': r.district_n,
                    'agency': r.agency_nam,
                    'basin': r.basin_name
                }
            })

        return jsonify({'type': 'FeatureCollection', 'features': features})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
@app.route('/api/recommend-category', methods=['POST'])
def api_recommend_category():
    """Enhanced API endpoint for category recommendations with scoring and alternatives."""
    try:
        data = request.get_json()

        # Extract user data
        user_data = {
            'roof_area': data.get('roof_area', 100),
            'open_space': data.get('open_space', 20),
            'household_size': data.get('household_size', 4)
        }

        # Extract location data
        location_data = {
            'Rainfall_mm': data.get('rainfall', 1000),
            'Soil_Type': data.get('soil_type', 'Loamy'),
            'Groundwater_Depth_m': data.get('gw_depth', 10),
            'Infiltration_Rate_mm_per_hr': data.get('infiltration_rate', 15)
        }

        # Extract user preferences
        user_preferences = {
            'complexity_preference': data.get('complexity_preference', 'balanced')
        }

        # Get enhanced recommendations
        from recommendations import get_category_recommendations_with_preferences
        result = get_category_recommendations_with_preferences(
            user_data, location_data, user_preferences
        )

        return jsonify(result)

    except Exception as e:
        return jsonify({
            'error': str(e),
            'recommended_category': {
                'id': 1,
                'name': 'Storage Tank Only',
                'description': 'Basic storage solution due to processing error',
                'confidence_score': 0,
                'recommendation_reason': 'Fallback due to error'
            }
        }), 500